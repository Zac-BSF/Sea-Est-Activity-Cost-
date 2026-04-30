"""
Process production Excel files into structured JSON for the Activity Cost Dashboard.

Usage:
    python scripts/process_excel.py "path/to/excel_file.xlsx"
    python scripts/process_excel.py "path/to/excel_file.xlsx" --append

The script reads all tabs (Skinner, Slicer for Skin on, Slicer for Skinless, Stripping),
cleans the data, calculates labor cost per finished lb, and outputs to data/production_data.json.

Use --append to add new data to the existing JSON (deduplicates by date+activity+lot+pallet).
"""

import sys
import os
import json
import re
from datetime import datetime, timedelta
from statistics import mean, median, stdev
import openpyxl
from openpyxl.cell.cell import MergedCell

LABOR_RATE = 22.00
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "production_data_v2.json")

# Weekly protein prices ($/lb) - weeks defined as Monday-Sunday
# Key: Monday date of the week
PROTEIN_PRICES = {
    "2026-03-09": {"skin_on": 6.27, "abf": 6.55, "coho": 5.45, "steelhead": 5.90, "sockeye": 10.47, "grouper": 10.55, "snapper": 7.97},
    "2026-03-16": {"skin_on": 6.27, "abf": 6.55, "coho": 5.45, "steelhead": 5.90, "sockeye": 10.47, "grouper": 10.55, "snapper": 7.97},
    "2026-03-23": {"skin_on": 6.37, "abf": 6.68, "coho": 5.45, "steelhead": 5.90, "sockeye": 10.47, "grouper": 10.55, "snapper": 7.97},
    "2026-03-30": {"skin_on": 6.37, "abf": 6.68, "coho": 5.45, "steelhead": 5.90, "sockeye": 10.47, "grouper": 10.55, "snapper": 7.97},
    "2026-04-06": {"skin_on": 6.37, "abf": 6.68, "coho": 5.45, "steelhead": 5.90, "sockeye": 10.47, "grouper": 10.55, "snapper": 7.97},
    "2026-04-13": {"skin_on": 6.46, "abf": 6.74, "coho": 5.45, "steelhead": 5.90, "sockeye": 10.47, "grouper": 10.55, "snapper": 7.97},
}


def get_week_monday(dt):
    """Get the Monday of the week containing dt (Mon-Sun week)."""
    days_since_monday = dt.weekday()  # Monday=0, Sunday=6
    monday = dt - timedelta(days=days_since_monday)
    return monday.strftime('%Y-%m-%d')


def get_protein_price(dt, activity, product_format):
    """Look up the raw protein price for a record based on its date and product."""
    monday = get_week_monday(dt)
    prices = PROTEIN_PRICES.get(monday)
    if not prices:
        return None

    fmt_lower = product_format.lower() if product_format else ""

    # Skinner
    if activity == "Skinner":
        if "abf" in fmt_lower:
            return prices["abf"]
        return prices["skin_on"]  # Conventional

    # Slicer Skin-on - all are conventional salmon being sliced
    if activity == "Slicer Skin-on":
        return prices["skin_on"]

    # Slicer Skinless
    if activity == "Slicer Skinless":
        if "abf" in fmt_lower:
            return prices["abf"]
        return prices["skin_on"]

    # Stripping - map by species
    if activity == "Stripping":
        if "coho" in fmt_lower:
            return prices["coho"]
        if "steelhead" in fmt_lower:
            return prices["steelhead"]
        if "sockeye" in fmt_lower:
            return prices["sockeye"]
        if "grouper" in fmt_lower:
            return prices["grouper"]
        if "snapper" in fmt_lower:
            return prices["snapper"]
        if "salmon" in fmt_lower or "skin on" in fmt_lower:
            return prices["skin_on"]
        return None

    return None


def classify_record(record):
    """Add classification (Fresh/Previously Frozen) and standardize product format."""
    activity = record["activity"]
    fmt = record["product_format"]

    # --- Classification: Fresh vs Previously Frozen ---
    # Fresh = Atlantic salmon (skinned, sliced, stripped skin-on salmon)
    # Previously Frozen = Grouper, Snapper, Steelhead, Coho, Sockeye
    previously_frozen_species = ("grouper", "snapper", "steelhead", "coho", "sockeye")
    fmt_lower = fmt.lower() if fmt else ""

    if activity == "Stripping" and any(sp in fmt_lower for sp in previously_frozen_species):
        record["classification"] = "Previously Frozen"
    else:
        record["classification"] = "Fresh"

    # --- Standardize product format names ---
    if activity == "Skinner":
        if fmt == "ABF":
            record["product_format"] = "2-4 lb Skin-On ABF Atlantic Salmon Fillets"
        else:  # Conventional
            record["product_format"] = "2-4 lb Skin-On Atlantic Salmon Fillets"

    elif activity == "Slicer Skin-on":
        if "3-4" in fmt:
            record["product_format"] = "3-4 lb Skin-On Atlantic Salmon Fillets"
        elif "2-3" in fmt:
            record["product_format"] = "2-3 lb Skin-On Atlantic Salmon Fillets"
        else:  # ungraded
            record["product_format"] = "2-4 lb Skin-On Atlantic Salmon Fillets"

    elif activity == "Slicer Skinless":
        if "ABF" in fmt or fmt == "ABF":
            record["product_format"] = "2-4 lb Skin-On ABF Atlantic Salmon Fillets"
        elif "From Skin-on (ABF)" in fmt:
            record["product_format"] = "2-4 lb Skin-On ABF Atlantic Salmon Fillets"
        elif "From Skin-on (Conventional)" in fmt:
            record["product_format"] = "2-4 lb Skin-On Atlantic Salmon Fillets"
        else:  # Conventional
            record["product_format"] = "2-4 lb Skin-On Atlantic Salmon Fillets"

    elif activity == "Stripping":
        if "skin on" in fmt_lower and "salmon" in fmt_lower:
            record["product_format"] = "2-4 lb Skin-On Atlantic Salmon Fillets"
        # Leave species names (Coho, Sockeye, Steelhead, Snapper, Grouper) as-is

    return record


def enrich_with_protein_cost(record):
    """Add protein cost, yield loss cost, and total cost fields to a record."""
    dt = datetime.strptime(record["date"], '%Y-%m-%d')
    price = get_protein_price(dt, record["activity"], record["product_format"])
    record["raw_protein_cost_per_lb"] = price

    if record["activity"] == "Stripping":
        # Stripping has no yield loss (lbs in = lbs out), just labor
        record["protein_cost_per_finished_lb"] = price
        record["yield_loss_cost_per_lb"] = 0
        if price and record.get("cost_per_finished_lb"):
            record["total_cost_per_finished_lb"] = round(price + record["cost_per_finished_lb"], 4)
        elif record.get("cost_per_finished_lb"):
            record["total_cost_per_finished_lb"] = record["cost_per_finished_lb"]
        else:
            record["total_cost_per_finished_lb"] = None
    elif price and record["yield_pct"] and record["yield_pct"] > 0:
        protein_cost_per_finished = price / (record["yield_pct"] / 100.0)
        yield_loss_cost = protein_cost_per_finished - price
        record["protein_cost_per_finished_lb"] = round(protein_cost_per_finished, 4)
        record["yield_loss_cost_per_lb"] = round(yield_loss_cost, 4)
        if record.get("cost_per_finished_lb"):
            record["total_cost_per_finished_lb"] = round(
                protein_cost_per_finished + record["cost_per_finished_lb"], 4
            )
        else:
            record["total_cost_per_finished_lb"] = None
    else:
        record["protein_cost_per_finished_lb"] = None
        record["yield_loss_cost_per_lb"] = None
        record["total_cost_per_finished_lb"] = None

    return record


TARGET_COSTS = {
    "2-4 lb Skin-On Atlantic Salmon Fillets": 7.52,       # Cost threshold for skinless conventional
    "2-4 lb Skin-On ABF Atlantic Salmon Fillets": 7.84,   # Cost threshold for skinless ABF
}


def compute_chained_costs(records):
    """
    Recompute costs with upstream activities flowing downstream:
      Stripping labor → Skinner input cost → Skinner output cost → Slicer Skinless input cost

    Also computes Production Spread = Sell KPI - total output cost.
    """
    # --- Step 1: Weekly avg stripping labor for Fresh Atlantic salmon ---
    strip_labor_by_week = {}
    for r in records:
        if r["activity"] != "Stripping" or r["classification"] != "Fresh":
            continue
        if not r.get("cost_per_finished_lb"):
            continue
        w = r["week"]
        strip_labor_by_week.setdefault(w, []).append(r["cost_per_finished_lb"])

    avg_strip_labor = {w: mean(vals) for w, vals in strip_labor_by_week.items()}

    # --- Step 2: Recompute Skinner records with upstream stripping cost ---
    skinner_output_by_week_product = {}

    for r in records:
        if r["activity"] != "Skinner":
            continue
        if not r.get("raw_protein_cost_per_lb") or not r.get("yield_pct"):
            continue

        raw_price = r["raw_protein_cost_per_lb"]
        strip_labor = avg_strip_labor.get(r["week"], 0)
        input_cost = raw_price + strip_labor

        yield_frac = r["yield_pct"] / 100.0
        yielded_input_cost = input_cost / yield_frac  # protein + strip through yield loss
        labor = r["cost_per_finished_lb"] or 0
        output_cost = yielded_input_cost + labor

        # Update the record
        r["upstream_strip_labor"] = round(strip_labor, 4)
        r["input_cost_per_lb"] = round(input_cost, 4)
        r["total_cost_per_finished_lb"] = round(output_cost, 4)
        r["yield_loss_cost_per_lb"] = round(yielded_input_cost - input_cost, 4)
        r["protein_cost_per_finished_lb"] = round(yielded_input_cost, 4)

        # KPI and Production Spread
        kpi = TARGET_COSTS.get(r["product_format"])
        if kpi:
            spread = kpi - output_cost
            r["target_cost"] = kpi
            r["production_spread_per_lb"] = round(spread, 4)
            r["extended_production_spread"] = round(spread * r["finished_lbs"], 2)
        else:
            r["target_cost"] = None
            r["production_spread_per_lb"] = None
            r["extended_production_spread"] = None

        # Track weekly skinner output cost by product for downstream use
        key = (r["week"], r["product_format"])
        skinner_output_by_week_product.setdefault(key, []).append(output_cost)

    avg_skinner_output = {k: mean(v) for k, v in skinner_output_by_week_product.items()}

    # --- Step 3: Recompute Slicer Skinless with Skinner output as input cost ---
    for r in records:
        if r["activity"] != "Slicer Skinless":
            continue
        if not r.get("yield_pct"):
            continue

        # Find the matching skinner output cost (same week, same product type)
        upstream_cost = avg_skinner_output.get((r["week"], r["product_format"]))
        if not upstream_cost:
            # Try matching by base product (ABF or conventional)
            fmt = r["product_format"]
            for key, val in avg_skinner_output.items():
                if key[0] == r["week"] and key[1] == fmt:
                    upstream_cost = val
                    break
            if not upstream_cost:
                # Fallback: use any skinner output for that week
                week_costs = [v for k, v in avg_skinner_output.items() if k[0] == r["week"]]
                if week_costs:
                    upstream_cost = mean(week_costs)

        if not upstream_cost:
            continue

        yield_frac = r["yield_pct"] / 100.0
        yielded_input_cost = upstream_cost / yield_frac
        labor = r["cost_per_finished_lb"] or 0
        output_cost = yielded_input_cost + labor

        r["input_cost_per_lb"] = round(upstream_cost, 4)
        r["total_cost_per_finished_lb"] = round(output_cost, 4)
        r["yield_loss_cost_per_lb"] = round(yielded_input_cost - upstream_cost, 4)
        r["protein_cost_per_finished_lb"] = round(yielded_input_cost, 4)
        r["raw_protein_cost_per_lb"] = round(upstream_cost, 4)  # Override: input is skinner output

        kpi = TARGET_COSTS.get(r["product_format"])
        if kpi:
            spread = kpi - output_cost
            r["target_cost"] = kpi
            r["production_spread_per_lb"] = round(spread, 4)
            r["extended_production_spread"] = round(spread * r["finished_lbs"], 2)

    # --- Step 4: Add KPI/spread to Slicer Skin-on records too (no upstream recompute needed) ---
    for r in records:
        if r["activity"] == "Slicer Skin-on":
            if not r.get("target_cost"):
                r["target_cost"] = None
                r["production_spread_per_lb"] = None
                r["extended_production_spread"] = None
        if r["activity"] == "Stripping":
            r["target_cost"] = None
            r["production_spread_per_lb"] = None
            r["extended_production_spread"] = None

    return records


def parse_time(t_str):
    """Parse a time string like '6:00 AM' or '6:00AM' into hours since midnight."""
    t_str = t_str.strip().upper()
    t_str = re.sub(r'\s+', ' ', t_str)
    for fmt in ('%I:%M %p', '%I:%M%p', '%H:%M'):
        try:
            t = datetime.strptime(t_str, fmt)
            return t.hour + t.minute / 60.0
        except ValueError:
            continue
    return None


def parse_labor_time(labor_str):
    """
    Parse labor time string into total hours worked.
    Handles multiple work segments: "11:55 AM - 12:25 PM - 1:15 PM - 1:26 PM"
    means two segments: 11:55-12:25 and 1:15-1:26.
    """
    if not labor_str or not isinstance(labor_str, str):
        return None

    labor_str = labor_str.strip()
    if not labor_str:
        return None

    break_minutes = 0
    break_match = re.search(r"(\d+)['\u2019]\s*BREAK", labor_str, re.IGNORECASE)
    if break_match:
        break_minutes = int(break_match.group(1))

    parts = re.split(r'\s*-\s*', labor_str)
    times = []
    for p in parts:
        p = p.strip()
        if not p or 'BREAK' in p.upper() or 'LUNCH' in p.upper():
            continue
        t = parse_time(p)
        if t is not None:
            times.append(t)

    if len(times) < 2:
        return None

    total_hours = 0
    for i in range(0, len(times) - 1, 2):
        start = times[i]
        end = times[i + 1]
        if end < start:
            end += 12
        diff = end - start
        if diff > 0:
            total_hours += diff

    if len(times) % 2 == 1:
        pass

    total_hours -= break_minutes / 60.0
    return max(total_hours, 0) if total_hours > 0 else None


def normalize_supplier(s):
    if not s or not isinstance(s, str):
        return None
    s = s.strip().upper()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('`', '').replace("'", '')
    if s in ('MULTIX', 'MULTI X', 'MULTI  X'):
        return 'Multi-X'
    if s in ('AQUA', 'AQUA`'):
        return 'AquaChile'
    if s == 'CERMAQ':
        return 'Cermaq'
    if s == 'BLUGLACIER':
        return 'BluGlacier'
    if s == 'TRAPANANDA':
        return 'Trapananda'
    return s


def get_week_label(dt):
    iso = dt.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def cell_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return None
    return cell.value


def process_skinner(ws):
    records = []
    current_date = None
    current_supplier = None
    current_people = None

    for row_idx in range(5, ws.max_row + 1):
        date_val = cell_value(ws, row_idx, 1)
        supplier_val = cell_value(ws, row_idx, 2)
        lot_val = cell_value(ws, row_idx, 3)
        pallet_val = cell_value(ws, row_idx, 4)
        incoming = safe_float(cell_value(ws, row_idx, 6))
        outgoing = safe_float(cell_value(ws, row_idx, 7))
        yield_val = safe_float(cell_value(ws, row_idx, 8))
        product_format = cell_value(ws, row_idx, 9)
        people_val = cell_value(ws, row_idx, 11)
        labor_val = cell_value(ws, row_idx, 12)

        if date_val and hasattr(date_val, 'strftime'):
            current_date = date_val
            current_people = None
        if supplier_val and isinstance(supplier_val, str) and supplier_val.strip():
            current_supplier = supplier_val
        if people_val is not None:
            try:
                current_people = int(people_val)
            except (ValueError, TypeError):
                pass

        if incoming is None or outgoing is None:
            continue
        if incoming <= 0 or outgoing <= 0:
            continue
        if current_date is None:
            continue

        hours = parse_labor_time(str(labor_val) if labor_val else None)
        if hours is None or current_people is None:
            continue

        fmt = str(product_format).strip() if product_format else ""
        if fmt.upper() == 'ABF':
            fmt = "ABF"
        else:
            fmt = "Conventional"

        total_labor_hours = current_people * hours
        labor_cost = total_labor_hours * LABOR_RATE
        cost_per_lb = labor_cost / outgoing if outgoing > 0 else None
        yield_pct = (outgoing / incoming * 100) if incoming > 0 else None

        records.append({
            "activity": "Skinner",
            "date": current_date.strftime('%Y-%m-%d'),
            "week": get_week_label(current_date),
            "supplier": normalize_supplier(current_supplier),
            "lot": str(lot_val).strip() if lot_val else None,
            "pallet": str(pallet_val).strip() if pallet_val else None,
            "product_format": fmt,
            "incoming_lbs": round(incoming, 2),
            "finished_lbs": round(outgoing, 2),
            "yield_pct": round(yield_pct, 2) if yield_pct else None,
            "people": current_people,
            "hours_worked": round(hours, 4),
            "total_labor_hours": round(total_labor_hours, 4),
            "labor_cost": round(labor_cost, 2),
            "cost_per_finished_lb": round(cost_per_lb, 4) if cost_per_lb else None
        })

    return records


def process_slicer_skin_on(ws):
    records = []
    current_date = None
    current_supplier = None
    current_people = None

    for row_idx in range(6, ws.max_row + 1):
        date_val = cell_value(ws, row_idx, 1)
        supplier_val = cell_value(ws, row_idx, 2)
        lot_val = cell_value(ws, row_idx, 3)
        pallet_val = cell_value(ws, row_idx, 4)
        incoming = safe_float(cell_value(ws, row_idx, 6))
        sides = safe_float(cell_value(ws, row_idx, 7)) or 0
        portions = safe_float(cell_value(ws, row_idx, 8)) or 0
        pesto = safe_float(cell_value(ws, row_idx, 9)) or 0
        pieces = safe_float(cell_value(ws, row_idx, 10)) or 0
        yield_val = safe_float(cell_value(ws, row_idx, 11))
        product_format = cell_value(ws, row_idx, 12)
        people_val = cell_value(ws, row_idx, 14)
        labor_val = cell_value(ws, row_idx, 15)

        if date_val and hasattr(date_val, 'strftime'):
            current_date = date_val
            current_people = None
        if supplier_val and isinstance(supplier_val, str) and supplier_val.strip():
            current_supplier = supplier_val
        if people_val is not None:
            try:
                current_people = int(people_val)
            except (ValueError, TypeError):
                pass

        if incoming is None or incoming <= 0:
            continue
        if current_date is None:
            continue

        total_output = sides + portions + pesto + pieces
        if total_output <= 0:
            continue

        hours = parse_labor_time(str(labor_val) if labor_val else None)
        if hours is None or current_people is None:
            continue

        fmt = str(product_format).strip() if product_format else ""
        if fmt in ('None', '', ' '):
            fmt = "Skin on (ungraded)"

        total_labor_hours = current_people * hours
        labor_cost = total_labor_hours * LABOR_RATE
        cost_per_lb = labor_cost / total_output if total_output > 0 else None
        yield_pct = (total_output / incoming * 100) if incoming > 0 else None

        records.append({
            "activity": "Slicer Skin-on",
            "date": current_date.strftime('%Y-%m-%d'),
            "week": get_week_label(current_date),
            "supplier": normalize_supplier(current_supplier),
            "lot": str(lot_val).strip() if lot_val else None,
            "pallet": str(pallet_val).strip() if pallet_val else None,
            "product_format": fmt,
            "incoming_lbs": round(incoming, 2),
            "finished_lbs": round(total_output, 2),
            "yield_pct": round(yield_pct, 2) if yield_pct else None,
            "people": current_people,
            "hours_worked": round(hours, 4),
            "total_labor_hours": round(total_labor_hours, 4),
            "labor_cost": round(labor_cost, 2),
            "cost_per_finished_lb": round(cost_per_lb, 4) if cost_per_lb else None
        })

    return records


def process_slicer_skinless(ws):
    records = []
    current_date = None
    current_supplier = None
    current_people = None

    for row_idx in range(6, ws.max_row + 1):
        date_val = cell_value(ws, row_idx, 1)
        supplier_val = cell_value(ws, row_idx, 2)
        lot_val = cell_value(ws, row_idx, 3)
        pallet_val = cell_value(ws, row_idx, 4)
        incoming = safe_float(cell_value(ws, row_idx, 5))
        skinless_out = safe_float(cell_value(ws, row_idx, 6)) or 0
        pieces_out = safe_float(cell_value(ws, row_idx, 7)) or 0
        yield_val = safe_float(cell_value(ws, row_idx, 8))
        product_format = cell_value(ws, row_idx, 9)
        people_val = cell_value(ws, row_idx, 11)
        labor_val = cell_value(ws, row_idx, 12)

        if date_val and hasattr(date_val, 'strftime'):
            current_date = date_val
            current_people = None
        elif date_val and isinstance(date_val, str):
            date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', str(date_val))
            if date_match:
                try:
                    current_date = datetime.strptime(date_match.group(1), '%m/%d/%Y')
                    current_people = None
                except ValueError:
                    pass

        if supplier_val and isinstance(supplier_val, str) and supplier_val.strip():
            current_supplier = supplier_val
        if people_val is not None:
            try:
                current_people = int(people_val)
            except (ValueError, TypeError):
                pass

        if incoming is None or incoming <= 0:
            continue
        total_output = skinless_out + pieces_out
        if total_output <= 0:
            continue
        if current_date is None:
            continue

        hours = parse_labor_time(str(labor_val) if labor_val else None)
        if hours is None or current_people is None:
            continue

        fmt = str(product_format).strip() if product_format else ""
        fmt = re.sub(r'\s+', ' ', fmt).strip()
        if fmt in ('None', ''):
            fmt = "Conventional"
        if fmt == 'From skin on':
            fmt = "From Skin-on (Conventional)"
        elif fmt == 'From skin on ABF':
            fmt = "From Skin-on (ABF)"

        total_labor_hours = current_people * hours
        labor_cost = total_labor_hours * LABOR_RATE
        cost_per_lb = labor_cost / total_output if total_output > 0 else None
        yield_pct = (total_output / incoming * 100) if incoming > 0 else None

        records.append({
            "activity": "Slicer Skinless",
            "date": current_date.strftime('%Y-%m-%d'),
            "week": get_week_label(current_date),
            "supplier": normalize_supplier(current_supplier),
            "lot": str(lot_val).strip() if lot_val else None,
            "pallet": str(pallet_val).strip() if pallet_val else None,
            "product_format": fmt,
            "incoming_lbs": round(incoming, 2),
            "finished_lbs": round(total_output, 2),
            "yield_pct": round(yield_pct, 2) if yield_pct else None,
            "people": current_people,
            "hours_worked": round(hours, 4),
            "total_labor_hours": round(total_labor_hours, 4),
            "labor_cost": round(labor_cost, 2),
            "cost_per_finished_lb": round(cost_per_lb, 4) if cost_per_lb else None
        })

    return records


def process_stripping(ws):
    records = []

    for row_idx in range(6, ws.max_row + 1):
        date_val = cell_value(ws, row_idx, 1)
        product_val = cell_value(ws, row_idx, 2)
        lbs_val = safe_float(cell_value(ws, row_idx, 3))
        people_val = cell_value(ws, row_idx, 4)
        labor_val = cell_value(ws, row_idx, 5)

        if not date_val or not hasattr(date_val, 'strftime'):
            continue
        if lbs_val is None or lbs_val <= 0:
            continue

        try:
            people = int(people_val)
        except (ValueError, TypeError):
            continue

        hours = parse_labor_time(str(labor_val) if labor_val else None)
        if hours is None:
            continue

        fmt = str(product_val).strip() if product_val else "Unknown"

        total_labor_hours = people * hours
        labor_cost = total_labor_hours * LABOR_RATE
        cost_per_lb = labor_cost / lbs_val if lbs_val > 0 else None

        records.append({
            "activity": "Stripping",
            "date": date_val.strftime('%Y-%m-%d'),
            "week": get_week_label(date_val),
            "supplier": None,
            "lot": None,
            "pallet": None,
            "product_format": fmt,
            "incoming_lbs": round(lbs_val, 2),
            "finished_lbs": round(lbs_val, 2),
            "yield_pct": None,
            "people": people,
            "hours_worked": round(hours, 4),
            "total_labor_hours": round(total_labor_hours, 4),
            "labor_cost": round(labor_cost, 2),
            "cost_per_finished_lb": round(cost_per_lb, 4) if cost_per_lb else None
        })

    return records


def compute_summary(records):
    groups = {}
    for r in records:
        key = f"{r['activity']}|{r['product_format']}"
        groups.setdefault(key, []).append(r)

    summary = {}
    for key, recs in groups.items():
        costs = [r['cost_per_finished_lb'] for r in recs if r['cost_per_finished_lb'] is not None and r['cost_per_finished_lb'] > 0]
        yields = [r['yield_pct'] for r in recs if r['yield_pct'] is not None]
        total_finished = sum(r['finished_lbs'] for r in recs)

        if costs:
            sorted_costs = sorted(costs)
            n = len(sorted_costs)
            p25_idx = int(n * 0.25)
            p75_idx = int(n * 0.75)

            total_costs = [r['total_cost_per_finished_lb'] for r in recs if r.get('total_cost_per_finished_lb')]
            yield_loss_costs = [r['yield_loss_cost_per_lb'] for r in recs if r.get('yield_loss_cost_per_lb')]

            summary[key] = {
                "count": len(recs),
                "avg_cost": round(mean(costs), 4),
                "median_cost": round(median(costs), 4),
                "min_cost": round(min(costs), 4),
                "max_cost": round(max(costs), 4),
                "p25_cost": round(sorted_costs[p25_idx], 4),
                "p75_cost": round(sorted_costs[min(p75_idx, n - 1)], 4),
                "std_cost": round(stdev(costs), 4) if len(costs) > 1 else 0,
                "avg_yield": round(mean(yields), 2) if yields else None,
                "avg_yield_loss_cost": round(mean(yield_loss_costs), 4) if yield_loss_costs else None,
                "avg_total_cost": round(mean(total_costs), 4) if total_costs else None,
                "total_finished_lbs": round(total_finished, 2)
            }

    return summary


def main():
    if len(sys.argv) < 2:
        print("Usage: python process_excel.py <excel_file> [--append]")
        sys.exit(1)

    excel_path = sys.argv[1]
    append_mode = '--append' in sys.argv

    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")
        sys.exit(1)

    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    all_records = []

    if 'Skinner' in wb.sheetnames:
        recs = process_skinner(wb['Skinner'])
        print(f"  Skinner: {len(recs)} records")
        all_records.extend(recs)

    if 'Slicer for Skin on ' in wb.sheetnames:
        recs = process_slicer_skin_on(wb['Slicer for Skin on '])
        print(f"  Slicer Skin-on: {len(recs)} records")
        all_records.extend(recs)

    if 'Slicer for Skinless' in wb.sheetnames:
        recs = process_slicer_skinless(wb['Slicer for Skinless'])
        print(f"  Slicer Skinless: {len(recs)} records")
        all_records.extend(recs)

    if 'Stripping' in wb.sheetnames:
        recs = process_stripping(wb['Stripping'])
        print(f"  Stripping: {len(recs)} records")
        all_records.extend(recs)

    wb.close()

    # Enrich all records with protein cost data
    # Enrich with protein cost first (needs original format names), then classify/rename
    all_records = [enrich_with_protein_cost(r) for r in all_records]
    all_records = [classify_record(r) for r in all_records]

    # Compute chained costs: stripping → skinning → slicing
    all_records = compute_chained_costs(all_records)

    # Rename activities for display
    ACTIVITY_NAMES = {
        "Skinner": "Skinning",
        "Slicer Skin-on": "Slicing - Skin-On Salmon",
        "Slicer Skinless": "Slicing - Skinless Salmon",
        "Stripping": "Stripping",
    }
    for r in all_records:
        r["activity"] = ACTIVITY_NAMES.get(r["activity"], r["activity"])

    if append_mode and os.path.exists(OUTPUT_PATH):
        with open(OUTPUT_PATH, 'r') as f:
            existing = json.load(f)
        existing_keys = set()
        for r in existing['records']:
            k = f"{r['date']}|{r['activity']}|{r.get('lot','')}|{r.get('pallet','')}"
            existing_keys.add(k)
        new_count = 0
        for r in all_records:
            k = f"{r['date']}|{r['activity']}|{r.get('lot','')}|{r.get('pallet','')}"
            if k not in existing_keys:
                existing['records'].append(r)
                new_count += 1
        all_records = existing['records']
        print(f"\nAppend mode: added {new_count} new records")

    summary = compute_summary(all_records)

    output = {
        "generated_at": datetime.now().isoformat(),
        "labor_rate": LABOR_RATE,
        "protein_prices": PROTEIN_PRICES,
        "source_file": os.path.basename(excel_path),
        "total_records": len(all_records),
        "records": sorted(all_records, key=lambda r: (r['date'], r['activity'])),
        "summary": summary
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\nTotal records: {len(all_records)}")
    print(f"Output: {OUTPUT_PATH}")
    print("\nSummary by Activity|Product:")
    for key, stats in sorted(summary.items()):
        print(f"  {key}: avg ${stats['avg_cost']:.4f}/lb, range ${stats['min_cost']:.4f}-${stats['max_cost']:.4f}, n={stats['count']}")


if __name__ == '__main__':
    main()
